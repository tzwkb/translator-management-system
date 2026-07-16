"""审计日志、导入导出接口。"""
import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..db import engine
from ..models import PO, AuditLog, LanguagePair, QualityScore, Translator
from ..schemas import POIn, TranslatorIn
from ..security import require_editor
from ..services import audit, resync_translator, svc_create_po

router = APIRouter(prefix="/api")

TR_COLS = ["id"] + list(Translator.EDITABLE) + list(Translator.DERIVED)   # 全 52 列

# 导入模板字段（人录入项）：(中文表头, 字段, 是否必填, 枚举可选值)
IMPORT_FIELDS = [
    ("姓名", "name", True, None),
    ("母语", "native_language", True, None),
    ("入库日期", "onboarding_date", True, None),
    ("邮箱", "email", False, None),
    ("微信", "wechat", False, None),
    ("所在地", "location", False, None),
    ("时区", "timezone", False, None),
    ("状态", "status", False, ["Active", "Dormant", "Blacklisted", "Probation"]),
    ("来源", "source", False, None),
    ("语言对", "language_pairs", False, None),
    ("擅长领域", "domains", False, None),
    ("文本类型", "text_types", False, None),
    ("CAT工具", "cat_tools", False, None),
    ("内部评级", "internal_rating", False, ["S", "A", "B", "C", "D"]),
    ("试译结果", "trial_result", False, ["Pass", "Fail", "Pending"]),
    ("当前项目", "current_project", False, None),
    ("角色", "role", False, ["翻译", "审校", "MTPE", "LQA"]),
    ("日产字数", "daily_output", False, None),
    ("是否双休", "weekend_off", False, ["是", "否"]),
    ("档期", "availability", False, ["空闲", "部分空闲", "满负荷"]),
    ("结算币种", "currency", False, ["CNY", "USD", "EUR"]),
    ("支付方式", "payment_method", False, None),
    ("开票类型", "invoice_type", False, None),
    ("扣税方式", "tax_deduction", False, None),
    ("合同状态", "contract_status", False, ["有效", "即将到期", "已过期", "无合同"]),
    ("合同到期", "contract_expiry", False, None),
    ("NDA签署", "nda_signed", False, ["是", "否"]),
    ("准时率", "punctuality_rate", False, None),
    ("响应速度", "responsiveness", False, ["快", "一般", "慢"]),
    ("合作评级", "cooperation_rating", False, ["优先合作", "正常合作", "谨慎合作", "停止合作"]),
    ("最近联系", "last_contact", False, None),
    ("备注", "remarks", False, None),
]
_HEADER_MAP = {lbl: f for lbl, f, _, _ in IMPORT_FIELDS}
_HEADER_MAP.update({f: f for _, f, _, _ in IMPORT_FIELDS})   # 英文字段名也认
_BOOL_FIELDS = {"weekend_off", "nda_signed"}
_DATE_FIELDS = {"onboarding_date", "contract_expiry", "last_contact"}
_PO_HEADER_MAP = {
    "translator_id": "translator_id", "译员ID": "translator_id", "译员id": "translator_id", "ID": "translator_id",
    "translator": "translator_name", "translator_name": "translator_name", "译员": "translator_name", "姓名": "translator_name",
    "settlement_month": "settlement_month", "结算月": "settlement_month", "月份": "settlement_month", "年月": "settlement_month",
    "project": "project", "项目": "project",
    "source_lang": "source_lang", "源语言": "source_lang", "源语": "source_lang",
    "target_lang": "target_lang", "目标语言": "target_lang", "目标语": "target_lang",
    "role": "role", "任务类型": "role", "类型": "role", "角色": "role",
    "word_count": "word_count", "字数": "word_count", "字数（字）": "word_count", "字符数": "word_count",
    "rate": "rate", "单价": "rate", "单价（/千字）": "rate", "费率": "rate",
    "currency": "currency", "币种": "currency",
    "status": "status", "状态": "status",
    "po_number": "po_number", "PO号": "po_number", "PO 号": "po_number", "po_number": "po_number",
    "remarks": "remarks", "备注": "remarks",
}


def _norm_header(c):
    return str(c).replace("（必填）", "").replace("*", "").strip() if c is not None else ""


def _norm_month_value(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m")
    s = str(v).strip()
    return s[:7] if len(s) >= 7 else s


def _blank(v):
    return v is None or (isinstance(v, str) and not v.strip())


def _find_translator_id(s: Session, raw_id, raw_name):
    if not _blank(raw_id):
        try:
            tid = int(raw_id)
        except (TypeError, ValueError):
            raise ValueError("译员ID不是整数") from None
        t = s.get(Translator, tid)
        if not t or t.deleted_at:
            raise ValueError("译员ID不存在")
        return tid
    if _blank(raw_name):
        raise ValueError("缺译员")
    name = str(raw_name).strip()
    rows = s.scalars(select(Translator).where(Translator.deleted_at.is_(None))).all()
    matches = [t for t in rows if t.name == name]
    if not matches:
        matches = [t for t in rows if t.name and (t.name in name or name in t.name)]
    if not matches:
        raise ValueError("译员不存在")
    if len(matches) > 1:
        raise ValueError("译员匹配不唯一")
    return matches[0].id


# ---------------- 审计 ----------------
@router.get("/audit")
def list_audit(who: str = Depends(require_editor)):
    with Session(engine) as s:
        rows = s.scalars(select(AuditLog).order_by(AuditLog.id.desc()).limit(100)).all()
        return [r.as_dict() for r in rows]


# ---------------- 导入导出 ----------------
@router.get("/export/translators")
def export_translators():
    from openpyxl import Workbook
    with Session(engine) as s:
        rows = s.scalars(select(Translator).where(Translator.deleted_at.is_(None)).order_by(Translator.id)).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "译员"
        ws.append(TR_COLS)
        for t in rows:
            d = t.as_dict()
            ws.append([d.get(c) for c in TR_COLS])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=translators.xlsx"})


@router.post("/import/translators")
def import_translators(file: UploadFile, who: str = Depends(require_editor)):
    """按导入模板（中文表头，也认英文字段名）批量建档；邮箱重复跳过；派生字段不导入。"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file.file.read()))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0}
    header = [_norm_header(c) for c in rows[0]]
    n, skipped, invalid_rows = 0, 0, []
    with Session(engine) as s:
        existing = {e for (e,) in s.execute(select(Translator.email).where(Translator.deleted_at.is_(None))).all() if e}
        for row_no, r in enumerate(rows[1:], start=2):
            data = {}
            for col, val in zip(header, r):
                f = _HEADER_MAP.get(col)
                if not f or val is None or val == "":
                    continue
                if f in _BOOL_FIELDS:
                    val = str(val).strip() in ("是", "true", "True", "1", "Y", "y")
                elif f in _DATE_FIELDS:
                    val = str(val)[:10]
                data[f] = val
            if not data.get("name"):
                continue
            data["name"] = str(data["name"])
            data.setdefault("status", "Active")
            try:
                clean = TranslatorIn(**data).model_dump()
            except ValidationError as e:
                invalid_rows.append({
                    "row": row_no,
                    "error": "；".join(f"{'.'.join(map(str, err['loc']))}: {err['msg']}" for err in e.errors(include_url=False)),
                })
                continue
            email = data.get("email")
            if email and email in existing:
                skipped += 1
                continue
            t = Translator(**clean)
            s.add(t)
            s.flush()
            # 语言对：把逗号分隔串拆成子表行
            lp_str = clean.get("language_pairs")
            if lp_str:
                for pair in str(lp_str).split(","):
                    p = pair.strip()
                    if "→" in p:
                        src, tgt = p.split("→", 1)
                        s.add(LanguagePair(translator_id=t.id, source_lang=src.strip(), target_lang=tgt.strip()))
            if email:
                existing.add(email)
            n += 1
        audit(s, who, "导入", "译员", None, f"{n} 条(跳过重复{skipped}，格式错误{len(invalid_rows)})")
        s.commit()
    return {"imported": n, "skipped_dup_email": skipped, "invalid_rows": invalid_rows}


@router.post("/import/po")
def import_po(file: UploadFile, who: str = Depends(require_editor)):
    """导入标准 PO/结算表；支持译员姓名或 ID，重复 PO 号跳过，错误行返回明细。"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file.file.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "skipped_dup_po": 0, "invalid_rows": []}
    header = [_norm_header(c) for c in rows[0]]
    imported, skipped_dup_po, invalid_rows = 0, 0, []
    with Session(engine) as s:
        existing_po = {p for (p,) in s.execute(select(PO.po_number)).all() if p}
        for row_no, r in enumerate(rows[1:], start=2):
            raw = {}
            for col, val in zip(header, r):
                f = _PO_HEADER_MAP.get(col)
                if not f or _blank(val):
                    continue
                if f == "settlement_month":
                    val = _norm_month_value(val)
                raw[f] = val
            if not raw:
                continue
            try:
                data = {k: v for k, v in raw.items() if k != "translator_name"}
                data["translator_id"] = _find_translator_id(s, raw.get("translator_id"), raw.get("translator_name"))
                if data.get("po_number") and str(data["po_number"]).strip() in existing_po:
                    skipped_dup_po += 1
                    continue
                if data.get("po_number") is not None:
                    data["po_number"] = str(data["po_number"]).strip()
                clean = POIn(**data).model_dump()
                p = svc_create_po(s, clean, who)
                if p.po_number:
                    existing_po.add(p.po_number)
                imported += 1
            except (ValidationError, ValueError, HTTPException) as e:
                if isinstance(e, ValidationError):
                    msg = "；".join(
                        f"{'.'.join(map(str, err['loc']))}: {err['msg']}"
                        for err in e.errors(include_url=False)
                    )
                elif isinstance(e, HTTPException):
                    msg = str(e.detail)
                else:
                    msg = str(e)
                invalid_rows.append({"row": row_no, "error": msg})
        audit(s, who, "导入", "PO", None, f"{imported} 条(跳过重复{skipped_dup_po}，格式错误{len(invalid_rows)})")
        s.commit()
    return {"imported": imported, "skipped_dup_po": skipped_dup_po, "invalid_rows": invalid_rows}


@router.post("/import/lqe")
def import_lqe(file: UploadFile, period: Optional[str] = None, who: str = Depends(require_editor)):
    """读 LQE 流程产出的 <label>_LQE报告.xlsx 的「汇总」sheet，按子表名匹配译员，写入质量记分卡。"""
    import re

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file.file.read()), data_only=True)
    data_rows, project = None, None
    for ws in wb.worksheets:
        allr = list(ws.iter_rows(values_only=True))
        for r in allr[:4]:
            for c in r:
                if isinstance(c, str) and "项目" in c:
                    mp = re.search(r"项目\s*([^\s]+)", c)
                    if mp:
                        project = mp.group(1)
        for idx, r in enumerate(allr):
            if r and str(r[0]).strip() == "子表":
                data_rows = allr[idx + 1:]
                break
        if data_rows:
            break
    if data_rows is None:
        raise HTTPException(400, "未找到 LQE 汇总表（缺『子表/SCORE』表头）")
    period = period or date.today().strftime("%Y-%m")
    imported, unmatched, affected = 0, [], set()
    with Session(engine) as s:
        trs = s.scalars(select(Translator).where(Translator.deleted_at.is_(None))).all()
        for r in data_rows:
            name = str(r[0]).strip() if r and r[0] is not None else ""
            if not name or name == "合计":
                continue
            t = next((x for x in trs if x.name and (x.name in name or name in x.name)), None)
            if not t:
                unmatched.append(name)
                continue
            g = lambda i: (r[i] if i < len(r) else None)
            errs = int(g(3) or 0)
            crit = int(g(4) or 0)
            s.add(QualityScore(translator_id=t.id, evaluation_period=period, project=project,
                               qa_type="LQE", score=g(5), critical_errors=crit,
                               minor_errors=max(0, errs - crit),
                               reviewer="LQE自动", feedback_notes=f"LQE导入 段数{g(1)} 词数{g(2)} STATUS{g(6)}"))
            affected.add(t.id)
            imported += 1
        for tid in affected:
            resync_translator(s, tid)
        audit(s, who, "导入", "LQE质量分", None, f"{imported}条 未匹配{len(unmatched)}")
        s.commit()
    return {"imported": imported, "unmatched": unmatched, "project": project, "period": period}
