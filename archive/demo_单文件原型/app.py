import base64
import hashlib
import hmac
import os
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from fastapi import Depends, FastAPI, Header, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import (Boolean, DateTime, ForeignKey, Integer, LargeBinary, Numeric,
                        String, Text, create_engine, func, select)
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column

HERE = Path(__file__).parent
engine = create_engine(f"sqlite:///{HERE / 'demo.db'}")
PO_STATUSES = ["未开票", "已开票待付", "已支付", "有争议"]

# ---------------- 字段级加密（AES-256-GCM）----------------
KEYFILE = HERE / ".key"
if KEYFILE.exists():
    KEY = KEYFILE.read_bytes()
else:
    KEY = AESGCM.generate_key(bit_length=256)
    KEYFILE.write_bytes(KEY)
_aes = AESGCM(KEY)


def enc(text):
    if not text:
        return None
    nonce = os.urandom(12)
    return nonce + _aes.encrypt(nonce, str(text).encode(), None)


def dec(blob):
    if not blob:
        return None
    return _aes.decrypt(blob[:12], blob[12:], None).decode()


def mask(text):
    if not text:
        return None
    t = str(text)
    return ("*" * max(0, len(t) - 4)) + t[-4:] if len(t) > 4 else "****"


# ---------------- 鉴权（首版免密码，token 按角色签名）----------------
AUTH_SECRET = os.urandom(16).hex()
USERS = {"资源端": ("editor", "资源端"), "财务": ("editor", "财务"), "boss": ("viewer", "boss"),
         "资源端Agent": ("agent", "资源端Agent")}


def make_token(role, name):
    payload = f"{role}|{name}"
    sig = hmac.new(AUTH_SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()[:16]
    return base64.urlsafe_b64encode(f"{payload}.{sig}".encode()).decode()


def parse_token(tok):
    if not tok:
        return None, None
    try:
        raw = base64.urlsafe_b64decode(tok.encode()).decode()
    except Exception:
        return None, None
    if "." not in raw:
        return None, None
    payload, sig = raw.rsplit(".", 1)
    good = hmac.new(AUTH_SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()[:16]
    if hmac.compare_digest(sig, good):
        role, name = payload.split("|", 1)
        return role, name
    return None, None


def _tok(authorization):
    if not authorization:
        return None
    return authorization.split(" ", 1)[1] if " " in authorization else authorization


def require_editor(authorization: str = Header(None)):
    role, name = parse_token(_tok(authorization))
    if role != "editor":
        raise HTTPException(403, "只读视角无编辑权限")
    return name


# ---------------- 模型 ----------------
class Base(DeclarativeBase):
    pass


class Translator(Base):
    __tablename__ = "translators"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(100))
    wechat: Mapped[Optional[str]] = mapped_column(String(100))
    email: Mapped[Optional[str]] = mapped_column(String(200))
    language_pairs: Mapped[Optional[str]] = mapped_column(String(200))
    translation_rate: Mapped[Optional[float]] = mapped_column(Numeric(10, 2))
    status: Mapped[str] = mapped_column(String(20), default="Active")
    internal_rating: Mapped[Optional[str]] = mapped_column(String(20))
    current_project: Mapped[Optional[str]] = mapped_column(String(200))
    availability: Mapped[Optional[str]] = mapped_column(String(50))
    contract_expiry: Mapped[Optional[str]] = mapped_column(String(20))
    last_contact: Mapped[Optional[str]] = mapped_column(String(20))
    remarks: Mapped[Optional[str]] = mapped_column(Text)
    deleted_at: Mapped[Optional[datetime]] = mapped_column(DateTime)

    def as_dict(self):
        r = self.translation_rate
        return {"id": self.id, "name": self.name, "wechat": self.wechat, "email": self.email,
                "language_pairs": self.language_pairs,
                "translation_rate": float(r) if r is not None else None, "status": self.status,
                "internal_rating": self.internal_rating, "current_project": self.current_project,
                "availability": self.availability, "contract_expiry": self.contract_expiry,
                "last_contact": self.last_contact, "remarks": self.remarks}


class RateChange(Base):
    __tablename__ = "rate_changes"
    id: Mapped[int] = mapped_column(primary_key=True)
    translator_id: Mapped[int] = mapped_column(ForeignKey("translators.id"))
    change_date: Mapped[str] = mapped_column(String(20))
    task_type: Mapped[Optional[str]] = mapped_column(String(50))
    original_rate: Mapped[Optional[float]] = mapped_column(Numeric(10, 2))
    new_rate: Mapped[Optional[float]] = mapped_column(Numeric(10, 2))
    negotiator: Mapped[Optional[str]] = mapped_column(String(100))
    reason: Mapped[Optional[str]] = mapped_column(String(100))
    remarks: Mapped[Optional[str]] = mapped_column(Text)

    def as_dict(self):
        return {"id": self.id, "translator_id": self.translator_id, "change_date": self.change_date,
                "task_type": self.task_type,
                "original_rate": float(self.original_rate) if self.original_rate is not None else None,
                "new_rate": float(self.new_rate) if self.new_rate is not None else None,
                "negotiator": self.negotiator, "reason": self.reason, "remarks": self.remarks}


class PO(Base):
    __tablename__ = "po_settlements"
    id: Mapped[int] = mapped_column(primary_key=True)
    translator_id: Mapped[int] = mapped_column(ForeignKey("translators.id"))
    settlement_month: Mapped[str] = mapped_column(String(7))
    project: Mapped[Optional[str]] = mapped_column(String(200))
    role: Mapped[Optional[str]] = mapped_column(String(50))
    word_count: Mapped[Optional[float]] = mapped_column(Numeric(12, 2))
    rate: Mapped[Optional[float]] = mapped_column(Numeric(10, 2))
    amount: Mapped[Optional[float]] = mapped_column(Numeric(12, 2))
    currency: Mapped[str] = mapped_column(String(10), default="CNY")
    status: Mapped[str] = mapped_column(String(20), default="未开票")
    po_number: Mapped[Optional[str]] = mapped_column(String(50))
    remarks: Mapped[Optional[str]] = mapped_column(Text)

    def as_dict(self):
        def f(x):
            return float(x) if x is not None else None
        return {"id": self.id, "translator_id": self.translator_id,
                "settlement_month": self.settlement_month, "project": self.project, "role": self.role,
                "word_count": f(self.word_count), "rate": f(self.rate), "amount": f(self.amount),
                "currency": self.currency, "status": self.status, "po_number": self.po_number,
                "remarks": self.remarks}


class Contract(Base):
    __tablename__ = "contracts"
    id: Mapped[int] = mapped_column(primary_key=True)
    translator_id: Mapped[int] = mapped_column(ForeignKey("translators.id"))
    contract_number: Mapped[Optional[str]] = mapped_column(String(100))
    contract_type: Mapped[Optional[str]] = mapped_column(String(50))
    sign_date: Mapped[Optional[str]] = mapped_column(String(20))
    expiry_date: Mapped[Optional[str]] = mapped_column(String(20))
    nda_signed: Mapped[bool] = mapped_column(Boolean, default=False)
    nda_expiry: Mapped[Optional[str]] = mapped_column(String(20))
    status: Mapped[Optional[str]] = mapped_column(String(20))
    remarks: Mapped[Optional[str]] = mapped_column(Text)

    def as_dict(self):
        return {"id": self.id, "translator_id": self.translator_id,
                "contract_number": self.contract_number, "contract_type": self.contract_type,
                "sign_date": self.sign_date, "expiry_date": self.expiry_date,
                "nda_signed": self.nda_signed, "nda_expiry": self.nda_expiry,
                "status": self.status, "remarks": self.remarks}


class QualityScore(Base):
    __tablename__ = "quality_scores"
    id: Mapped[int] = mapped_column(primary_key=True)
    translator_id: Mapped[int] = mapped_column(ForeignKey("translators.id"))
    evaluation_period: Mapped[Optional[str]] = mapped_column(String(20))
    project: Mapped[Optional[str]] = mapped_column(String(200))
    qa_type: Mapped[Optional[str]] = mapped_column(String(50))
    score: Mapped[Optional[float]] = mapped_column(Numeric(5, 2))
    critical_errors: Mapped[int] = mapped_column(Integer, default=0)
    major_errors: Mapped[int] = mapped_column(Integer, default=0)
    minor_errors: Mapped[int] = mapped_column(Integer, default=0)
    reviewer: Mapped[Optional[str]] = mapped_column(String(100))
    feedback_notes: Mapped[Optional[str]] = mapped_column(Text)

    def as_dict(self):
        return {"id": self.id, "translator_id": self.translator_id,
                "evaluation_period": self.evaluation_period, "project": self.project,
                "qa_type": self.qa_type, "score": float(self.score) if self.score is not None else None,
                "critical_errors": self.critical_errors, "major_errors": self.major_errors,
                "minor_errors": self.minor_errors, "reviewer": self.reviewer,
                "feedback_notes": self.feedback_notes}


class Complaint(Base):
    __tablename__ = "complaints"
    id: Mapped[int] = mapped_column(primary_key=True)
    translator_id: Mapped[int] = mapped_column(ForeignKey("translators.id"))
    date: Mapped[Optional[str]] = mapped_column(String(20))
    project: Mapped[Optional[str]] = mapped_column(String(200))
    complaint_type: Mapped[Optional[str]] = mapped_column(String(50))
    severity: Mapped[Optional[str]] = mapped_column(String(20))
    deduction_amount: Mapped[Optional[float]] = mapped_column(Numeric(12, 2))
    resolution: Mapped[Optional[str]] = mapped_column(String(50))
    remarks: Mapped[Optional[str]] = mapped_column(Text)

    def as_dict(self):
        return {"id": self.id, "translator_id": self.translator_id, "date": self.date,
                "project": self.project, "complaint_type": self.complaint_type,
                "severity": self.severity,
                "deduction_amount": float(self.deduction_amount) if self.deduction_amount is not None else None,
                "resolution": self.resolution, "remarks": self.remarks}


class Capacity(Base):
    __tablename__ = "capacity_allocations"
    id: Mapped[int] = mapped_column(primary_key=True)
    translator_id: Mapped[int] = mapped_column(ForeignKey("translators.id"))
    period_year: Mapped[int] = mapped_column(Integer)
    period_month: Mapped[int] = mapped_column(Integer)
    week_no: Mapped[int] = mapped_column(Integer)
    project: Mapped[Optional[str]] = mapped_column(String(200))
    occupancy_pct: Mapped[int] = mapped_column(Integer, default=0)

    def as_dict(self):
        return {"id": self.id, "translator_id": self.translator_id, "period_year": self.period_year,
                "period_month": self.period_month, "week_no": self.week_no, "project": self.project,
                "occupancy_pct": self.occupancy_pct}


class PaymentInfo(Base):
    __tablename__ = "payment_infos"
    translator_id: Mapped[int] = mapped_column(ForeignKey("translators.id"), primary_key=True)
    currency: Mapped[Optional[str]] = mapped_column(String(10))
    bank_name: Mapped[Optional[str]] = mapped_column(String(200))
    bank_account_enc: Mapped[Optional[bytes]] = mapped_column(LargeBinary)
    id_card_enc: Mapped[Optional[bytes]] = mapped_column(LargeBinary)
    payee_name: Mapped[Optional[str]] = mapped_column(String(100))
    supports_wechat: Mapped[bool] = mapped_column(Boolean, default=False)
    remarks: Mapped[Optional[str]] = mapped_column(Text)

    def as_dict(self, reveal=False):
        ba = dec(self.bank_account_enc)
        ic = dec(self.id_card_enc)
        return {"translator_id": self.translator_id, "currency": self.currency,
                "bank_name": self.bank_name,
                "bank_account": ba if reveal else mask(ba),
                "id_card": ic if reveal else mask(ic),
                "payee_name": self.payee_name, "supports_wechat": self.supports_wechat,
                "remarks": self.remarks}


class AuditLog(Base):
    __tablename__ = "audit_logs"
    id: Mapped[int] = mapped_column(primary_key=True)
    user: Mapped[Optional[str]] = mapped_column(String(50))
    action: Mapped[Optional[str]] = mapped_column(String(20))
    entity: Mapped[Optional[str]] = mapped_column(String(50))
    entity_id: Mapped[Optional[int]] = mapped_column(Integer)
    detail: Mapped[Optional[str]] = mapped_column(Text)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.now)

    def as_dict(self):
        return {"id": self.id, "user": self.user, "action": self.action, "entity": self.entity,
                "entity_id": self.entity_id, "detail": self.detail,
                "created_at": self.created_at.strftime("%Y-%m-%d %H:%M:%S") if self.created_at else None}


class PendingChange(Base):
    __tablename__ = "pending_changes"
    id: Mapped[int] = mapped_column(primary_key=True)
    created_by: Mapped[Optional[str]] = mapped_column(String(50))
    kind: Mapped[Optional[str]] = mapped_column(String(30))
    translator_id: Mapped[Optional[int]] = mapped_column(Integer)
    payload: Mapped[Optional[str]] = mapped_column(Text)
    status: Mapped[str] = mapped_column(String(20), default="pending")
    reviewed_by: Mapped[Optional[str]] = mapped_column(String(50))
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.now)


def audit(s, user, action, entity, eid, detail=""):
    s.add(AuditLog(user=user, action=action, entity=entity, entity_id=eid, detail=detail))


# ---------------- Pydantic ----------------
class TranslatorIn(BaseModel):
    name: str
    wechat: Optional[str] = None
    email: Optional[str] = None
    language_pairs: Optional[str] = None
    translation_rate: Optional[float] = None
    status: str = "Active"
    internal_rating: Optional[str] = None
    current_project: Optional[str] = None
    availability: Optional[str] = None
    contract_expiry: Optional[str] = None
    last_contact: Optional[str] = None
    remarks: Optional[str] = None


class RateChangeIn(BaseModel):
    change_date: str
    task_type: Optional[str] = None
    original_rate: Optional[float] = None
    new_rate: Optional[float] = None
    negotiator: Optional[str] = None
    reason: Optional[str] = None
    remarks: Optional[str] = None


class POIn(BaseModel):
    translator_id: int
    settlement_month: str
    project: Optional[str] = None
    role: Optional[str] = None
    word_count: Optional[float] = None
    rate: Optional[float] = None
    currency: str = "CNY"
    status: str = "未开票"
    po_number: Optional[str] = None
    remarks: Optional[str] = None


class StatusIn(BaseModel):
    status: str


class ContractIn(BaseModel):
    contract_number: Optional[str] = None
    contract_type: Optional[str] = None
    sign_date: Optional[str] = None
    expiry_date: Optional[str] = None
    nda_signed: bool = False
    nda_expiry: Optional[str] = None
    status: Optional[str] = None
    remarks: Optional[str] = None


class QualityIn(BaseModel):
    evaluation_period: Optional[str] = None
    project: Optional[str] = None
    qa_type: Optional[str] = None
    score: Optional[float] = None
    critical_errors: int = 0
    major_errors: int = 0
    minor_errors: int = 0
    reviewer: Optional[str] = None
    feedback_notes: Optional[str] = None


class ComplaintIn(BaseModel):
    date: Optional[str] = None
    project: Optional[str] = None
    complaint_type: Optional[str] = None
    severity: Optional[str] = None
    deduction_amount: Optional[float] = None
    resolution: Optional[str] = None
    remarks: Optional[str] = None


class CapacityIn(BaseModel):
    period_year: int
    period_month: int
    week_no: int
    project: Optional[str] = None
    occupancy_pct: int = 0


class PaymentIn(BaseModel):
    currency: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    id_card: Optional[str] = None
    payee_name: Optional[str] = None
    supports_wechat: bool = False
    remarks: Optional[str] = None


class LoginIn(BaseModel):
    user: str


Base.metadata.create_all(engine)


def seed():
    with Session(engine) as s:
        if s.scalar(select(func.count()).select_from(Translator)):
            return
        ts = [
            Translator(name="张明", wechat="zhangming_wx", email="zm@example.com", language_pairs="ZH→EN",
                       translation_rate=180, status="Active", internal_rating="A", current_project="燕云",
                       availability="部分空闲", contract_expiry="2026-12-31", last_contact="2026-06-20"),
            Translator(name="李娜", wechat="lina_t", email="lina@example.com", language_pairs="ZH→JA",
                       translation_rate=220, status="Active", internal_rating="S", current_project="洛克王国",
                       availability="满负荷", contract_expiry="2026-09-15", last_contact="2026-06-24"),
            Translator(name="王伟", email="ww@example.com", language_pairs="ZH→KO", translation_rate=150,
                       status="Dormant", internal_rating="B", availability="空闲",
                       contract_expiry="2026-07-10", last_contact="2026-03-01"),
            Translator(name="陈静", wechat="chenjing", email="cj@example.com", language_pairs="EN→ZH",
                       translation_rate=130, status="Probation", internal_rating="C", current_project="诺诺",
                       availability="部分空闲", contract_expiry="2027-01-20", last_contact="2026-06-18"),
        ]
        s.add_all(ts)
        s.commit()
        i = {t.name: t.id for t in ts}
        s.add_all([
            RateChange(translator_id=i["张明"], change_date="2026-05-10", task_type="翻译", original_rate=160,
                       new_rate=180, negotiator="Raven", reason="年度调价", remarks="量稳定，上调"),
            RateChange(translator_id=i["李娜"], change_date="2026-06-01", task_type="翻译", original_rate=200,
                       new_rate=220, negotiator="Raven", reason="译员主动要求", remarks="资深，原价低于市场"),
            RateChange(translator_id=i["李娜"], change_date="2026-06-10", task_type="LQE", original_rate=None,
                       new_rate=90, negotiator="Raven", reason="新增任务类型", remarks="LQE 单价，与 LQA 不同"),
        ])
        s.add_all([
            PO(translator_id=i["张明"], settlement_month="2026-06", project="燕云", role="翻译", word_count=8.0,
               rate=180, amount=1440, currency="CNY", status="已开票待付", po_number="PO-202606-001"),
            PO(translator_id=i["李娜"], settlement_month="2026-06", project="洛克王国", role="翻译", word_count=6.5,
               rate=220, amount=1430, currency="CNY", status="已支付", po_number="PO-202606-002"),
            PO(translator_id=i["陈静"], settlement_month="2026-06", project="诺诺", role="翻译", word_count=5.0,
               rate=130, amount=650, currency="CNY", status="未开票", po_number="PO-202606-003"),
            PO(translator_id=i["张明"], settlement_month="2026-06", project="海外SLG", role="翻译", word_count=3.0,
               rate=30, amount=90, currency="USD", status="已开票待付", po_number="PO-202606-004"),
            PO(translator_id=i["陈静"], settlement_month="2026-06", project="诺诺", role="审校", word_count=5.0,
               rate=80, amount=410, currency="CNY", status="未开票", po_number="PO-202606-005"),
        ])
        s.add_all([
            Contract(translator_id=i["张明"], contract_number="C-2025-011", contract_type="框架协议",
                     sign_date="2025-01-01", expiry_date="2026-12-31", nda_signed=True, nda_expiry="2027-12-31",
                     status="有效"),
            Contract(translator_id=i["王伟"], contract_number="C-2025-027", contract_type="单项目合同",
                     sign_date="2025-07-10", expiry_date="2026-07-10", nda_signed=True, nda_expiry="2026-07-10",
                     status="即将到期"),
        ])
        s.add_all([
            QualityScore(translator_id=i["张明"], evaluation_period="2026-05", project="燕云", qa_type="终检",
                         score=94, critical_errors=0, major_errors=1, minor_errors=3, reviewer="Raven",
                         feedback_notes="术语稳，个别润色"),
            QualityScore(translator_id=i["李娜"], evaluation_period="2026-06", project="洛克王国", qa_type="终检",
                         score=97, critical_errors=0, major_errors=0, minor_errors=2, reviewer="Raven",
                         feedback_notes="高质量"),
        ])
        s.add_all([
            Complaint(translator_id=i["陈静"], date="2026-06-12", project="诺诺", complaint_type="质量不达标",
                      severity="一般", deduction_amount=200, resolution="已扣款", remarks="客户指出风格不一致"),
        ])
        s.add_all([
            Capacity(translator_id=i["张明"], period_year=2026, period_month=6, week_no=1, project="燕云", occupancy_pct=60),
            Capacity(translator_id=i["张明"], period_year=2026, period_month=6, week_no=2, project="燕云", occupancy_pct=60),
            Capacity(translator_id=i["李娜"], period_year=2026, period_month=6, week_no=1, project="洛克王国", occupancy_pct=100),
        ])
        s.add(PaymentInfo(translator_id=i["张明"], currency="CNY", bank_name="招商银行",
                          bank_account_enc=enc("6225888812345678"), id_card_enc=enc("310101199001011234"),
                          payee_name="张明", supports_wechat=True))
        s.commit()


seed()

app = FastAPI(title="译员管理系统 Demo")
app.mount("/assets", StaticFiles(directory=HERE / "assets"), name="assets")


def names_of(s):
    return {t.id: t.name for t in s.scalars(select(Translator)).all()}


def get_translator(s, tid):
    t = s.get(Translator, tid)
    if not t or t.deleted_at:
        raise HTTPException(404, "译员不存在")
    return t


def require_writer(authorization: str = Header(None)):
    role, name = parse_token(_tok(authorization))
    if role not in ("editor", "agent"):
        raise HTTPException(403, "无编辑权限")
    return role, name


def svc_add_rate_change(s, tid, d, who):
    t = get_translator(s, tid)
    s.add(RateChange(translator_id=tid, change_date=d.get("change_date"), task_type=d.get("task_type"),
                     original_rate=d.get("original_rate"), new_rate=d.get("new_rate"),
                     negotiator=d.get("negotiator"), reason=d.get("reason"), remarks=d.get("remarks")))
    if d.get("task_type") == "翻译" and d.get("new_rate") is not None:
        t.translation_rate = d["new_rate"]
    audit(s, who, "新增", "报价变更", tid, f"{t.name} {d.get('task_type')} {d.get('new_rate')}")


def svc_create_po(s, d, who):
    amt = round((d.get("word_count") or 0) * (d.get("rate") or 0), 2)
    p = PO(translator_id=d["translator_id"], settlement_month=d["settlement_month"], project=d.get("project"),
           role=d.get("role"), word_count=d.get("word_count"), rate=d.get("rate"), amount=amt,
           currency=d.get("currency", "CNY"), status=d.get("status", "未开票"),
           po_number=d.get("po_number"), remarks=d.get("remarks"))
    s.add(p)
    s.flush()
    audit(s, who, "新增", "PO", p.id, f"{d['settlement_month']} {amt}")
    return p


def svc_set_po_status(s, pid, status, who):
    p = s.get(PO, pid)
    if not p:
        raise HTTPException(404, "PO 不存在")
    p.status = status
    audit(s, who, "改状态", "PO", pid, status)


def make_pending(s, name, kind, tid, payload):
    import json
    pc = PendingChange(created_by=name, kind=kind, translator_id=tid,
                       payload=json.dumps(payload, ensure_ascii=False))
    s.add(pc)
    audit(s, name, "提交建议", kind, tid, "")
    s.commit()
    s.refresh(pc)
    return {"pending": True, "pending_id": pc.id, "msg": "已提交待人工审核（钱相关需人工确认）"}


@app.get("/")
def index():
    return FileResponse(HERE / "index.html", headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


@app.post("/api/login")
def login(body: LoginIn):
    if body.user not in USERS:
        raise HTTPException(404, "用户不存在")
    role, name = USERS[body.user]
    return {"token": make_token(role, name), "role": role, "name": name}


@app.get("/api/me")
def me(authorization: str = Header(None)):
    role, name = parse_token(_tok(authorization))
    return {"role": role, "name": name}


@app.get("/api/overview")
def overview():
    today = date.today().isoformat()
    soon = (date.today() + timedelta(days=30)).isoformat()
    with Session(engine) as s:
        rows = s.scalars(select(Translator).where(Translator.deleted_at.is_(None))).all()
        by, expiring = {}, 0
        for t in rows:
            by[t.status] = by.get(t.status, 0) + 1
            if t.contract_expiry and today <= t.contract_expiry <= soon:
                expiring += 1
        return {"total": len(rows), "active": by.get("Active", 0), "dormant": by.get("Dormant", 0),
                "blacklisted": by.get("Blacklisted", 0), "probation": by.get("Probation", 0),
                "expiring_30d": expiring}


# ---- 译员 ----
@app.get("/api/translators")
def list_translators():
    with Session(engine) as s:
        rows = s.scalars(select(Translator).where(Translator.deleted_at.is_(None)).order_by(Translator.id)).all()
        return [r.as_dict() for r in rows]


@app.post("/api/translators")
def create_translator(body: TranslatorIn, who: str = Depends(require_editor)):
    with Session(engine) as s:
        if body.email and s.scalar(select(Translator).where(Translator.email == body.email, Translator.deleted_at.is_(None))):
            raise HTTPException(400, "邮箱已存在")
        t = Translator(**body.model_dump())
        s.add(t)
        s.flush()
        audit(s, who, "新增", "译员", t.id, t.name)
        s.commit()
        s.refresh(t)
        return t.as_dict()


@app.put("/api/translators/{tid}")
def update_translator(tid: int, body: TranslatorIn, who: str = Depends(require_editor)):
    with Session(engine) as s:
        t = get_translator(s, tid)
        if body.email and s.scalar(select(Translator).where(Translator.email == body.email, Translator.id != tid, Translator.deleted_at.is_(None))):
            raise HTTPException(400, "邮箱已存在")
        for k, v in body.model_dump().items():
            setattr(t, k, v)
        audit(s, who, "编辑", "译员", tid, t.name)
        s.commit()
        s.refresh(t)
        return t.as_dict()


@app.delete("/api/translators/{tid}")
def delete_translator(tid: int, who: str = Depends(require_editor)):
    with Session(engine) as s:
        t = get_translator(s, tid)
        t.deleted_at = datetime.now()
        audit(s, who, "删除", "译员", tid, t.name)
        s.commit()
        return {"ok": True}


# ---- 报价变更 ----
@app.get("/api/translators/{tid}/rate-changes")
def list_rate_changes(tid: int):
    with Session(engine) as s:
        rows = s.scalars(select(RateChange).where(RateChange.translator_id == tid)
                         .order_by(RateChange.change_date.desc(), RateChange.id.desc())).all()
        return [r.as_dict() for r in rows]


@app.post("/api/translators/{tid}/rate-changes")
def add_rate_change(tid: int, body: RateChangeIn, w=Depends(require_writer)):
    role, name = w
    with Session(engine) as s:
        get_translator(s, tid)
        d = body.model_dump()
        if role == "agent":
            return make_pending(s, name, "rate_change", tid, d)
        svc_add_rate_change(s, tid, d, name)
        s.commit()
        return {"ok": True}


# ---- 质量记分卡 ----
@app.get("/api/translators/{tid}/quality")
def list_quality(tid: int):
    with Session(engine) as s:
        rows = s.scalars(select(QualityScore).where(QualityScore.translator_id == tid)
                         .order_by(QualityScore.evaluation_period.desc(), QualityScore.id.desc())).all()
        return [r.as_dict() for r in rows]


@app.post("/api/translators/{tid}/quality")
def add_quality(tid: int, body: QualityIn, who: str = Depends(require_editor)):
    with Session(engine) as s:
        get_translator(s, tid)
        s.add(QualityScore(translator_id=tid, **body.model_dump()))
        audit(s, who, "新增", "质量记分", tid, f"{body.evaluation_period} {body.score}")
        s.commit()
        return {"ok": True}


# ---- 合同 ----
@app.get("/api/translators/{tid}/contracts")
def list_contracts(tid: int):
    with Session(engine) as s:
        rows = s.scalars(select(Contract).where(Contract.translator_id == tid).order_by(Contract.id)).all()
        return [r.as_dict() for r in rows]


@app.post("/api/translators/{tid}/contracts")
def add_contract(tid: int, body: ContractIn, who: str = Depends(require_editor)):
    with Session(engine) as s:
        get_translator(s, tid)
        s.add(Contract(translator_id=tid, **body.model_dump()))
        audit(s, who, "新增", "合同", tid, body.contract_number or "")
        s.commit()
        return {"ok": True}


# ---- 客诉 ----
@app.get("/api/translators/{tid}/complaints")
def list_complaints(tid: int):
    with Session(engine) as s:
        rows = s.scalars(select(Complaint).where(Complaint.translator_id == tid)
                         .order_by(Complaint.date.desc(), Complaint.id.desc())).all()
        return [r.as_dict() for r in rows]


@app.post("/api/translators/{tid}/complaints")
def add_complaint(tid: int, body: ComplaintIn, who: str = Depends(require_editor)):
    with Session(engine) as s:
        get_translator(s, tid)
        s.add(Complaint(translator_id=tid, **body.model_dump()))
        audit(s, who, "新增", "客诉", tid, f"{body.complaint_type} {body.severity}")
        s.commit()
        return {"ok": True}


# ---- 产能（占用 + 可用率；利用率口径待 PM，不计算）----
@app.get("/api/translators/{tid}/capacity")
def list_capacity(tid: int):
    with Session(engine) as s:
        rows = s.scalars(select(Capacity).where(Capacity.translator_id == tid)
                         .order_by(Capacity.period_year, Capacity.period_month, Capacity.week_no)).all()
        by_week = {}
        for r in rows:
            by_week.setdefault((r.period_year, r.period_month, r.week_no), 0)
            by_week[(r.period_year, r.period_month, r.week_no)] += r.occupancy_pct
        weeks = [{"period": f"{y}-{mo:02d} W{w}", "occupancy": occ, "available": 100 - occ}
                 for (y, mo, w), occ in sorted(by_week.items())]
        return {"rows": [r.as_dict() for r in rows], "weeks": weeks}


@app.post("/api/translators/{tid}/capacity")
def add_capacity(tid: int, body: CapacityIn, w=Depends(require_writer)):
    role, name = w
    with Session(engine) as s:
        get_translator(s, tid)
        s.add(Capacity(translator_id=tid, **body.model_dump()))
        audit(s, name, "新增", "产能", tid, f"{body.period_year}-{body.period_month} W{body.week_no} {body.occupancy_pct}%")
        s.commit()
        return {"ok": True}


# ---- 支付信息（敏感字段加密 + 脱敏）----
@app.get("/api/translators/{tid}/payment")
def get_payment(tid: int):
    with Session(engine) as s:
        p = s.get(PaymentInfo, tid)
        return p.as_dict() if p else None


@app.get("/api/translators/{tid}/payment/reveal")
def reveal_payment(tid: int, who: str = Depends(require_editor)):
    with Session(engine) as s:
        p = s.get(PaymentInfo, tid)
        if not p:
            raise HTTPException(404, "无支付信息")
        audit(s, who, "查看明文", "支付信息", tid, "")
        s.commit()
        return p.as_dict(reveal=True)


@app.put("/api/translators/{tid}/payment")
def put_payment(tid: int, body: PaymentIn, who: str = Depends(require_editor)):
    with Session(engine) as s:
        get_translator(s, tid)
        p = s.get(PaymentInfo, tid) or PaymentInfo(translator_id=tid)
        p.currency = body.currency
        p.bank_name = body.bank_name
        p.payee_name = body.payee_name
        p.supports_wechat = body.supports_wechat
        p.remarks = body.remarks
        if body.bank_account:
            p.bank_account_enc = enc(body.bank_account)
        if body.id_card:
            p.id_card_enc = enc(body.id_card)
        s.add(p)
        audit(s, who, "保存", "支付信息", tid, "")
        s.commit()
        return p.as_dict()


# ---- PO ----
@app.get("/api/po")
def list_po(month: Optional[str] = None, status: Optional[str] = None):
    with Session(engine) as s:
        q = select(PO)
        if month:
            q = q.where(PO.settlement_month == month)
        if status:
            q = q.where(PO.status == status)
        rows = s.scalars(q.order_by(PO.settlement_month.desc(), PO.id)).all()
        nm = names_of(s)
        out = []
        for r in rows:
            d = r.as_dict()
            d["translator_name"] = nm.get(r.translator_id, "?")
            exp = round((r.word_count or 0) * (r.rate or 0), 2)
            d["expected_amount"] = exp
            d["amount_ok"] = abs((r.amount or 0) - exp) <= 0.02
            out.append(d)
        return out


@app.get("/api/po/summary")
def po_summary(month: str):
    with Session(engine) as s:
        rows = s.scalars(select(PO).where(PO.settlement_month == month)).all()
        cur = {}
        for r in rows:
            c = cur.setdefault(r.currency or "?", {"unpaid": 0.0, "paid": 0.0})
            amt = float(r.amount or 0)
            if r.status == "已支付":
                c["paid"] += amt
            elif r.status in ("未开票", "已开票待付"):
                c["unpaid"] += amt
        return {"month": month, "by_currency": cur}


@app.post("/api/po")
def create_po(body: POIn, w=Depends(require_writer)):
    role, name = w
    with Session(engine) as s:
        d = body.model_dump()
        if role == "agent":
            return make_pending(s, name, "po", d["translator_id"], d)
        p = svc_create_po(s, d, name)
        s.commit()
        s.refresh(p)
        return p.as_dict()


@app.put("/api/po/{pid}/status")
def set_po_status(pid: int, body: StatusIn, w=Depends(require_writer)):
    role, name = w
    if body.status not in PO_STATUSES:
        raise HTTPException(400, "状态非法")
    with Session(engine) as s:
        if role == "agent":
            return make_pending(s, name, "po_status", None, {"pid": pid, "status": body.status})
        svc_set_po_status(s, pid, body.status, name)
        s.commit()
        return {"ok": True}


# ---- 审计日志 ----
@app.get("/api/audit")
def list_audit(who: str = Depends(require_editor)):
    with Session(engine) as s:
        rows = s.scalars(select(AuditLog).order_by(AuditLog.id.desc()).limit(100)).all()
        return [r.as_dict() for r in rows]


# ---- 导入导出 ----
TR_COLS = ["id", "name", "wechat", "email", "language_pairs", "translation_rate", "status",
           "internal_rating", "current_project", "availability", "contract_expiry", "last_contact", "remarks"]


@app.get("/api/export/translators")
def export_translators():
    import io
    from openpyxl import Workbook
    with Session(engine) as s:
        rows = s.scalars(select(Translator).where(Translator.deleted_at.is_(None)).order_by(Translator.id)).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "译员"
        ws.append(TR_COLS)
        for t in rows:
            d = t.as_dict()
            ws.append([d.get(c) for c in TR_COLS])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=translators.xlsx"})


@app.post("/api/import/translators")
def import_translators(file: UploadFile, who: str = Depends(require_editor)):
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file.file.read()))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0}
    header = [str(c) if c is not None else "" for c in rows[0]]
    n = 0
    with Session(engine) as s:
        existing = {e for (e,) in s.execute(select(Translator.email).where(Translator.deleted_at.is_(None))).all() if e}
        for r in rows[1:]:
            rec = dict(zip(header, r))
            name = rec.get("name")
            if not name:
                continue
            email = rec.get("email")
            if email and email in existing:
                continue
            s.add(Translator(name=str(name), email=email, wechat=rec.get("wechat"),
                             language_pairs=rec.get("language_pairs"),
                             translation_rate=rec.get("translation_rate"),
                             status=rec.get("status") or "Active",
                             internal_rating=rec.get("internal_rating"),
                             current_project=rec.get("current_project"),
                             contract_expiry=str(rec["contract_expiry"]) if rec.get("contract_expiry") else None,
                             remarks=rec.get("remarks")))
            if email:
                existing.add(email)
            n += 1
        audit(s, who, "导入", "译员", None, f"{n} 条")
        s.commit()
    return {"imported": n}


@app.post("/api/import/lqe")
def import_lqe(file: UploadFile, period: Optional[str] = None, who: str = Depends(require_editor)):
    """读 LQE 流程产出的 <label>_LQE报告.xlsx 的「汇总」sheet，按子表名匹配译员，写入质量记分卡。"""
    import io
    import re
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file.file.read()), data_only=True)
    data_rows, project = None, None
    for ws in wb.worksheets:
        allr = list(ws.iter_rows(values_only=True))
        for r in allr[:4]:
            for c in r:
                if isinstance(c, str) and "项目" in c:
                    mp = re.search(r"项目\s*([^\s]+)", c)
                    if mp:
                        project = mp.group(1)
        for idx, r in enumerate(allr):
            if r and str(r[0]).strip() == "子表":
                data_rows = allr[idx + 1:]
                break
        if data_rows:
            break
    if data_rows is None:
        raise HTTPException(400, "未找到 LQE 汇总表（缺『子表/SCORE』表头）")
    period = period or date.today().strftime("%Y-%m")
    imported, unmatched = 0, []
    with Session(engine) as s:
        trs = s.scalars(select(Translator).where(Translator.deleted_at.is_(None))).all()
        for r in data_rows:
            name = str(r[0]).strip() if r and r[0] is not None else ""
            if not name or name == "合计":
                continue
            t = next((x for x in trs if x.name and (x.name in name or name in x.name)), None)
            if not t:
                unmatched.append(name)
                continue
            g = lambda i: (r[i] if i < len(r) else None)
            errs = int(g(3) or 0)
            crit = int(g(4) or 0)
            s.add(QualityScore(translator_id=t.id, evaluation_period=period, project=project,
                               qa_type="LQE", score=g(5), critical_errors=crit,
                               minor_errors=max(0, errs - crit),
                               reviewer="LQE自动", feedback_notes=f"LQE导入 段数{g(1)} 词数{g(2)} STATUS{g(6)}"))
            imported += 1
        audit(s, who, "导入", "LQE质量分", None, f"{imported}条 未匹配{len(unmatched)}")
        s.commit()
    return {"imported": imported, "unmatched": unmatched, "project": project, "period": period}


@app.get("/api/pending")
def list_pending(who: str = Depends(require_editor)):
    import json
    with Session(engine) as s:
        rows = s.scalars(select(PendingChange).where(PendingChange.status == "pending")
                         .order_by(PendingChange.id.desc())).all()
        nm = names_of(s)
        return [{"id": r.id, "created_by": r.created_by, "kind": r.kind, "translator_id": r.translator_id,
                 "translator_name": nm.get(r.translator_id, "") if r.translator_id else "",
                 "payload": json.loads(r.payload),
                 "created_at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else None} for r in rows]


@app.post("/api/pending/{pcid}/approve")
def approve_pending(pcid: int, who: str = Depends(require_editor)):
    import json
    with Session(engine) as s:
        pc = s.get(PendingChange, pcid)
        if not pc or pc.status != "pending":
            raise HTTPException(404, "无此待审或已处理")
        d = json.loads(pc.payload)
        if pc.kind == "rate_change":
            svc_add_rate_change(s, pc.translator_id, d, f"{who}<approve>")
        elif pc.kind == "po":
            svc_create_po(s, d, f"{who}<approve>")
        elif pc.kind == "po_status":
            svc_set_po_status(s, d["pid"], d["status"], f"{who}<approve>")
        pc.status = "approved"
        pc.reviewed_by = who
        audit(s, who, "批准", "待审", pcid, pc.kind)
        s.commit()
        return {"ok": True}


@app.post("/api/pending/{pcid}/reject")
def reject_pending(pcid: int, who: str = Depends(require_editor)):
    with Session(engine) as s:
        pc = s.get(PendingChange, pcid)
        if not pc or pc.status != "pending":
            raise HTTPException(404, "无此待审或已处理")
        pc.status = "rejected"
        pc.reviewed_by = who
        audit(s, who, "驳回", "待审", pcid, pc.kind)
        s.commit()
        return {"ok": True}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
